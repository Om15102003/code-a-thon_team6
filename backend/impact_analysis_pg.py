import psycopg2
from psycopg2.extras import DictCursor
import pandas as pd
from tabulate import tabulate
from typing import Dict, List
from datetime import datetime, date
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment

EXCEL_FILE = 'impact_analysis.xlsx'

def connect_to_db():
    """Establish database connection"""
    return psycopg2.connect(
        dbname="employee_system",
        user="postgres",
        password="Omg$15102003",  # Replace with your PostgreSQL password
        host="localhost"
    )

def fetch_project_impacts() -> List[Dict]:
    """Fetch impact analysis data from PostgreSQL"""
    conn = connect_to_db()
    cur = conn.cursor(cursor_factory=DictCursor)
    
    try:
        # Get all projects
        cur.execute("""
            SELECT 
                "ImpactID",
                "Name",
                "Up_to_15_min_F", "Up_to_15_min_O", "Up_to_15_min_L", "Up_to_15_min_B",
                "Up_to_30_min_F", "Up_to_30_min_O", "Up_to_30_min_L", "Up_to_30_min_B",
                "Up_to_1_Hour_F", "Up_to_1_Hour_O", "Up_to_1_Hour_L", "Up_to_1_Hour_B",
                "Up_to_4_Hours_F", "Up_to_4_Hours_O", "Up_to_4_Hours_L", "Up_to_4_Hours_B",
                "Up_to_8_Hours_F", "Up_to_8_Hours_O", "Up_to_8_Hours_L", "Up_to_8_Hours_B",
                "Up_to_12_Hours_F", "Up_to_12_Hours_O", "Up_to_12_Hours_L", "Up_to_12_Hours_B",
                "Up_to_24_Hours_F", "Up_to_24_Hours_O", "Up_to_24_Hours_L", "Up_to_24_Hours_B",
                "Up_to_48_Hours_F", "Up_to_48_Hours_O", "Up_to_48_Hours_L", "Up_to_48_Hours_B",
                "Up_to_72_Hours_F", "Up_to_72_Hours_O", "Up_to_72_Hours_L", "Up_to_72_Hours_B",
                "Up_to_1_Week_F", "Up_to_1_Week_O", "Up_to_1_Week_L", "Up_to_1_Week_B",
                "AnalysisDate"
            FROM 
                "ImpactAnalysis"
            ORDER BY 
                "AnalysisDate" DESC
        """)
        
        results = cur.fetchall()
        
        processes = []
        for row in results:
            process = {
                'Process Name': row['Name'],
                'Type of Impact': ['FImpact', 'OCImpact', 'LDImpact', 'BImpact', 'HighestImpact'],
                'Up to 15 min': [
                    row['Up_to_15_min_F'],
                    row['Up_to_15_min_O'],
                    row['Up_to_15_min_L'],
                    row['Up_to_15_min_B'],
                    max(row['Up_to_15_min_F'], row['Up_to_15_min_O'], 
                        row['Up_to_15_min_L'], row['Up_to_15_min_B'])
                ],
                'Up to 30 min': [
                    row['Up_to_30_min_F'],
                    row['Up_to_30_min_O'],
                    row['Up_to_30_min_L'],
                    row['Up_to_30_min_B'],
                    max(row['Up_to_30_min_F'], row['Up_to_30_min_O'],
                        row['Up_to_30_min_L'], row['Up_to_30_min_B'])
                ],
                'Up to 1 Hour': [
                    row['Up_to_1_Hour_F'],
                    row['Up_to_1_Hour_O'],
                    row['Up_to_1_Hour_L'],
                    row['Up_to_1_Hour_B'],
                    max(row['Up_to_1_Hour_F'], row['Up_to_1_Hour_O'],
                        row['Up_to_1_Hour_L'], row['Up_to_1_Hour_B'])
                ],
                'Up to 4 Hours': [
                    row['Up_to_4_Hours_F'],
                    row['Up_to_4_Hours_O'],
                    row['Up_to_4_Hours_L'],
                    row['Up_to_4_Hours_B'],
                    max(row['Up_to_4_Hours_F'], row['Up_to_4_Hours_O'],
                        row['Up_to_4_Hours_L'], row['Up_to_4_Hours_B'])
                ],
                'Up to 8 Hours': [
                    row['Up_to_8_Hours_F'],
                    row['Up_to_8_Hours_O'],
                    row['Up_to_8_Hours_L'],
                    row['Up_to_8_Hours_B'],
                    max(row['Up_to_8_Hours_F'], row['Up_to_8_Hours_O'],
                        row['Up_to_8_Hours_L'], row['Up_to_8_Hours_B'])
                ],
                'Up to 12 Hours': [
                    row['Up_to_12_Hours_F'],
                    row['Up_to_12_Hours_O'],
                    row['Up_to_12_Hours_L'],
                    row['Up_to_12_Hours_B'],
                    max(row['Up_to_12_Hours_F'], row['Up_to_12_Hours_O'],
                        row['Up_to_12_Hours_L'], row['Up_to_12_Hours_B'])
                ],
                'Up to 24 Hours': [
                    row['Up_to_24_Hours_F'],
                    row['Up_to_24_Hours_O'],
                    row['Up_to_24_Hours_L'],
                    row['Up_to_24_Hours_B'],
                    max(row['Up_to_24_Hours_F'], row['Up_to_24_Hours_O'],
                        row['Up_to_24_Hours_L'], row['Up_to_24_Hours_B'])
                ],
                'Up to 48 Hours': [
                    row['Up_to_48_Hours_F'],
                    row['Up_to_48_Hours_O'],
                    row['Up_to_48_Hours_L'],
                    row['Up_to_48_Hours_B'],
                    max(row['Up_to_48_Hours_F'], row['Up_to_48_Hours_O'],
                        row['Up_to_48_Hours_L'], row['Up_to_48_Hours_B'])
                ],
                'Up to 72 Hours': [
                    row['Up_to_72_Hours_F'],
                    row['Up_to_72_Hours_O'],
                    row['Up_to_72_Hours_L'],
                    row['Up_to_72_Hours_B'],
                    max(row['Up_to_72_Hours_F'], row['Up_to_72_Hours_O'],
                        row['Up_to_72_Hours_L'], row['Up_to_72_Hours_B'])
                ],
                'Up to 1 Week': [
                    row['Up_to_1_Week_F'],
                    row['Up_to_1_Week_O'],
                    row['Up_to_1_Week_L'],
                    row['Up_to_1_Week_B'],
                    max(row['Up_to_1_Week_F'], row['Up_to_1_Week_O'],
                        row['Up_to_1_Week_L'], row['Up_to_1_Week_B'])
                ],
                'AnalysisDate': row['AnalysisDate']
            }
            processes.append(process)
        
        return processes
        
    finally:
        cur.close()
        conn.close()

def export_to_excel(processes: List[Dict]):
    """Export project data to Excel with matrix format"""
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        # Combine all processes into one DataFrame
        all_rows = []
        for process in processes:
            # Add process rows
            process_rows = []
            for i in range(len(process['Type of Impact'])):
                row_data = {
                    'Process Name': process['Process Name'] if i == 0 else '',
                    'Type of Impact': process['Type of Impact'][i]
                }
                # Add time columns
                for col in process.keys():
                    if col.startswith('Up to'):
                        row_data[col] = process[col][i]
                process_rows.append(row_data)
            all_rows.extend(process_rows)
            # Add blank row between processes
            all_rows.append({key: '' for key in all_rows[0].keys()})
        
        # Create DataFrame
        df = pd.DataFrame(all_rows)
        
        # Write to Excel
        df.to_excel(writer, sheet_name='Impact Analysis', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Impact Analysis']
        
        # Define styles
        header_fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 30  # Process Name
        worksheet.column_dimensions['B'].width = 15  # Type of Impact
        for col in range(3, 13):  # Time columns (C to L)
            worksheet.column_dimensions[chr(64 + col)].width = 12
        
        # Add borders and colors to all cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                # Skip header row and non-numeric cells
                if cell.row > 1 and isinstance(cell.value, (int, float)):
                    # Color based on value
                    if cell.value == 0:
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
                    elif cell.value == 1:
                        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Light Green
                    elif cell.value == 2:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
                    elif cell.value == 3:
                        cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange
                    elif cell.value >= 4:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
    
    print(f"Data exported to {EXCEL_FILE}")

def display_project_impacts(export=False):
    """Display impact analysis and optionally export to Excel"""
    try:
        processes = fetch_project_impacts()
        
        print("\nProject Impact Analysis Report")
        print("=" * 80)
        
        for process in processes:
            print(f"\nProcess: {process['Process Name']}")
            print(f"Analysis Date: {process['AnalysisDate'].strftime('%Y-%m-%d')}")
            print("-" * 60)
            
            # Create impact table
            impact_table = []
            for i, impact_type in enumerate(process['Type of Impact']):
                row = [impact_type]
                for col in process.keys():
                    if col.startswith('Up to'):
                        row.append(process[col][i])
                impact_table.append(row)
            
            print(tabulate(
                impact_table,
                headers=['Impact Type'] + [col for col in process.keys() if col.startswith('Up to')],
                tablefmt="grid",
                stralign="center"
            ))
            print()
        
        if export:
            export_to_excel(processes)
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    # Add command line argument handling if needed
    display_project_impacts(export=True)  # Set to True to also export to Excel 